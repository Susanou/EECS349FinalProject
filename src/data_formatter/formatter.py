import openpyxl

def create_workbook(title):
    wb = openpyxl.Workbook()
    wb.save(title)
    sheet = wb.active
    sheet.title = 'Threads'
    wb.create_sheet('Comments')
    __write_threads_data_header(sheet)
    sheet = wb['Comments']
    __write_comments_data_header(sheet)
    wb.save('Project T1_Template.xlsx')

    return wb


def __write_comments_data_header(sheet):
    sheet['A1'] = 'Thread_id'
    sheet['B1'] = 'Comment_link'
    sheet['B2'] = 'Specific page of the comment'
    sheet['C1'] = 'Floor_number'
    sheet['D1'] = 'User name'
    sheet['E1'] = 'Trade'
    sheet['E2'] = '[1] if the comment indicates the trade'
    sheet['F1'] = 'Review'
    sheet['F2'] = 'Sentiment of review (if the comment includes review)'
    sheet['G1'] = 'Q&A'
    sheet['G2'] = '[1] if the comment includes a Q&A'
    sheet['H1'] = 'Content'


def __write_threads_data_header(sheet):
    sheet['A1'] = 'Thread_id'
    sheet['B1'] = 'Thread_link'
    sheet['C1'] = 'Market'
    sheet['D1'] = 'Vendor name'
    sheet['E1'] = 'Product/service name'
    sheet['F1'] = 'Replies'
    sheet['G1'] = 'Views'
    sheet['H1'] = 'Category'
    sheet['H2'] = 'Crimeware or CaaS'
    sheet['I1'] = 'Price'
    sheet['J1'] = 'Unit'
    sheet['K1'] = 'Payment method'
    sheet['K2'] = '1:PayPal 2:CreditCard 3:Cryptocurrency 4: Others'

def write_data_points(sheet, data_tuple, row_num):
    true_row = str(row_num + 2)    # allow for the space buffer of the header
    for i, char in enumerate(__char_range('A', chr(ord('A') + len(data_tuple) - 1))):
        sheet[char + true_row] = data_tuple[i]


def __char_range(c1, c2):
    for c in range(ord(c1), ord(c2) + 1):
        yield chr(c)

if __name__ == '__main__':
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Data Analysis'
    __write_comments_data_header(sheet)
    data = ('111111111111', 'myfavoritehackerforum.org', 'cringe', 'ninja', 'bitchmodo', 'fornite', 'cringe')
    write_data_points(sheet, data, 1)
    wb.save('t1_g14_data.xlsx')